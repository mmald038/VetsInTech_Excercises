# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)

"""
# EASY MODE

# import the appropriate modules (you have 3)

import requests
import json
import openpyxl

character_url = "https://rickandmortyapi.com/api/character"
# set up a workbook and worksheet titled "Rick and Morty Characters"

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"

# # assign a variable 'data' with the returned GET request

response = requests.get(character_url)
data = response.json()


# create the appropriate headers in openpyxl for all of the keys for a single character

headers = list(data["results"][0].keys())

# print(headers)

# loop through all of the 'results' of the data to populate the rows and columns for each character
# Note: due to the headers, the rows need to be offset by one!

for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num, value=header)

for row_num, character in enumerate(data["results"], start=2):  # Offset by 1 for headers
    for col_num, key in enumerate(headers, start=1):
        ws.cell(row=row_num, column=col_num, value=str(character[key]))  # Convert values to string

# Save the workbook
wb.save("C:/Users/moise/Downloads/Data Analysis/VetsInTech/Python Fundamentals/Week 4/w3d3 Rick and Morty characters.xlsx")

"""
# MEDIUM MODE
import requests
import json
import openpyxl
from openpyxl import Workbook

# Create a new workbook and worksheets
wb = Workbook()
ws_characters = wb.active
ws_characters.title = "Rick and Morty Characters"

# Properly create new sheets and assign them to variables
ws_locations = wb.create_sheet(title="Rick and Morty Locations")
ws_episodes = wb.create_sheet(title="Rick and Morty Episodes")

# API URLs
character_url = "https://rickandmortyapi.com/api/character"
location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"

# Function to fetch and populate data into a worksheet
def fetch_and_populate(url, worksheet):
    response = requests.get(url)
    data = response.json()
    
    if "results" not in data:
        print(f"Error fetching data from {url}")
        return
    
    # Extract headers
    headers = list(data["results"][0].keys())

    # Write headers to the first row
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Populate rows with data
    for row_num, item in enumerate(data["results"], start=2):
        for col_num, key in enumerate(headers, start=1):
            worksheet.cell(row=row_num, column=col_num, value=str(item[key]))  # Convert values to string

# Fetch and populate data for characters, locations, and episodes
fetch_and_populate(character_url, ws_characters)
fetch_and_populate(location_url, ws_locations)
fetch_and_populate(episode_url, ws_episodes)

# Save the workbook
wb.save("C:/Users/moise/Downloads/Data Analysis/VetsInTech/Python Fundamentals/Week 4/w3d3 Rick and Morty Characters & Location & Episodes.xlsx")

# Print success message
print("Excel file created: Rick_and_Morty_Data.xlsx")


"""

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)






# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)


# wb.save("./spreadsheets/exercise.xlsx")
"""